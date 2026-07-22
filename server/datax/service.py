from __future__ import annotations

import ast
import csv
import hashlib
import json
import math
import re
import threading
import time
from datetime import date, datetime
from itertools import chain
from pathlib import Path
from typing import Any, Iterable

from .models import (
    DataSourceSnapshot,
    DataXImportJob,
    DataXProject,
    DataXResultArtifact,
    IndicatorDefinition,
    IndicatorFilter,
    IndicatorProposal,
    IndicatorVersion,
    ModelField,
    SemanticEntity,
    SemanticJoin,
    SemanticModel,
)
from .store import (
    DataXConflictError,
    DataXNotFoundError,
    DataXStore,
    DataXValidationError,
)


MAX_SOURCE_BYTES = 50 * 1024 * 1024
MAX_SOURCE_ROWS = 1_000_000
MAX_QUERY_ROWS = 500
ALLOWED_SOURCE_SUFFIXES = {".csv": "csv", ".xlsx": "xlsx", ".parquet": "parquet"}
ALLOWED_FORMULA_NODES = (
    ast.Expression,
    ast.BinOp,
    ast.UnaryOp,
    ast.Add,
    ast.Sub,
    ast.Mult,
    ast.Div,
    ast.USub,
    ast.UAdd,
    ast.Name,
    ast.Load,
    ast.Constant,
)


class DataXService:
    def __init__(self, store: DataXStore) -> None:
        self.store = store
        self._import_lock = threading.RLock()

    # Projects and immutable source snapshots
    def create_project(self, *, name: str, description: str = "") -> DataXProject:
        title = name.strip()
        if not title:
            raise DataXValidationError("Project name is required.")
        now = self.store.now()
        item = DataXProject(
            project_id=self.store.new_id("datax_project"),
            name=title[:160],
            description=description.strip()[:1000],
            created_at=now,
            updated_at=now,
        )
        return self.store.insert("projects", item, "project_id")  # type: ignore[return-value]

    def list_projects(self) -> list[DataXProject]:
        return sorted(
            self.store.list("projects", DataXProject),
            key=lambda item: (-item.updated_at, item.project_id),
        )

    def get_project(self, project_id: str) -> DataXProject:
        return self.store.require("projects", project_id, DataXProject)

    def update_project(
        self,
        project_id: str,
        *,
        revision: int,
        name: str | None = None,
        description: str | None = None,
        status: str | None = None,
    ) -> DataXProject:
        item = self.get_project(project_id)
        _check_revision(item.revision, revision)
        values = item.model_dump()
        if name is not None:
            if not name.strip():
                raise DataXValidationError("Project name is required.")
            values["name"] = name.strip()[:160]
        if description is not None:
            values["description"] = description.strip()[:1000]
        if status is not None:
            if status not in {"active", "archived"}:
                raise DataXValidationError("Invalid project status.")
            values["status"] = status
        values["revision"] += 1
        values["updated_at"] = self.store.now()
        updated = DataXProject.model_validate(values)
        return self.store.replace("projects", updated, "project_id")  # type: ignore[return-value]

    def import_source(self, project_id: str, *, file_name: str, content: bytes) -> DataXImportJob:
        job = self.create_import_job(project_id, file_name=file_name, content=content)
        return self.run_import_job(job.job_id)

    def create_import_job(
        self, project_id: str, *, file_name: str, content: bytes
    ) -> DataXImportJob:
        self.get_project(project_id)
        if len(content) > MAX_SOURCE_BYTES:
            raise DataXValidationError("Data source exceeds the 50 MB limit.")
        safe_name = Path(file_name).name
        if safe_name != file_name or not safe_name:
            raise DataXValidationError("Invalid source file name.")
        suffix = Path(safe_name).suffix.lower()
        file_type = ALLOWED_SOURCE_SUFFIXES.get(suffix)
        if file_type is None:
            raise DataXValidationError("Only CSV, XLSX, and Parquet files are supported.")
        digest = hashlib.sha256(content).hexdigest()
        existing = next(
            (
                source
                for source in self.list_sources(project_id)
                if source.content_sha256 == digest and source.status == "ready"
            ),
            None,
        )
        if existing is not None:
            now = self.store.now()
            job = DataXImportJob(
                job_id=self.store.new_id("datax_import"),
                project_id=project_id,
                source_id=existing.source_id,
                status="ready",
                attempt_count=0,
                created_at=now,
                updated_at=now,
                completed_at=now,
            )
            return self.store.insert("import_jobs", job, "job_id")  # type: ignore[return-value]

        now = self.store.now()
        source_id = self.store.new_id("datax_source")
        source = DataSourceSnapshot(
            source_id=source_id,
            project_id=project_id,
            name=Path(safe_name).stem[:160],
            file_name=safe_name[:240],
            file_type=file_type,  # type: ignore[arg-type]
            content_sha256=digest,
            byte_size=len(content),
            table_name=f"source_{source_id.replace('-', '_')}",
            created_at=now,
            updated_at=now,
        )
        job = DataXImportJob(
            job_id=self.store.new_id("datax_import"),
            project_id=project_id,
            source_id=source_id,
            created_at=now,
            updated_at=now,
        )
        self.store.insert("sources", source, "source_id")
        self.store.insert("import_jobs", job, "job_id")
        path = self.store.source_file_path(project_id, digest, suffix)
        self.store.write_source(path, content)
        return job

    def run_import_job(self, job_id: str, *, recover_stale_process: bool = False) -> DataXImportJob:
        with self._import_lock:
            job = self.store.require("import_jobs", job_id, DataXImportJob)
            if job.status == "ready":
                return job
            now = self.store.now()
            if (
                job.status == "processing"
                and job.lease_token
                and job.lease_expires_at is not None
                and job.lease_expires_at > now
                and not recover_stale_process
            ):
                return job
            source = self.store.require("sources", job.source_id, DataSourceSnapshot)
            lease = self.store.new_id("lease")
            job = job.model_copy(
                update={
                    "status": "processing",
                    "attempt_count": job.attempt_count + 1,
                    "lease_token": lease,
                    "lease_expires_at": now + 300,
                    "updated_at": now,
                    "error": "",
                }
            )
            self.store.replace("import_jobs", job, "job_id")
            source = source.model_copy(
                update={"status": "processing", "updated_at": now, "error": ""}
            )
            self.store.replace("sources", source, "source_id")
        try:
            source = self._load_snapshot(source)
        except Exception as exc:
            message = _safe_error(exc)
            failed_at = self.store.now()
            source = source.model_copy(
                update={"status": "failed", "error": message, "updated_at": failed_at}
            )
            job = job.model_copy(
                update={
                    "status": "failed",
                    "error": message,
                    "lease_token": None,
                    "lease_expires_at": None,
                    "updated_at": failed_at,
                    "completed_at": failed_at,
                }
            )
            self.store.replace("sources", source, "source_id")
            self.store.replace("import_jobs", job, "job_id")
            return job
        completed_at = self.store.now()
        self.store.replace("sources", source, "source_id")
        job = job.model_copy(
            update={
                "status": "ready",
                "lease_token": None,
                "lease_expires_at": None,
                "updated_at": completed_at,
                "completed_at": completed_at,
            }
        )
        return self.store.replace("import_jobs", job, "job_id")  # type: ignore[return-value]

    def recover_import_jobs(self) -> int:
        recovered = 0
        for job in self.store.list("import_jobs", DataXImportJob):
            if job.status in {"pending", "processing"}:
                self.run_import_job(job.job_id, recover_stale_process=True)
                recovered += 1
        return recovered

    def list_sources(self, project_id: str) -> list[DataSourceSnapshot]:
        self.get_project(project_id)
        return sorted(
            [item for item in self.store.list("sources", DataSourceSnapshot) if item.project_id == project_id],
            key=lambda item: (item.created_at, item.source_id),
        )

    def get_import_job(self, job_id: str) -> DataXImportJob:
        return self.store.require("import_jobs", job_id, DataXImportJob)

    def _load_snapshot(self, source: DataSourceSnapshot) -> DataSourceSnapshot:
        import duckdb

        suffix = "." + source.file_type
        path = self.store.source_file_path(source.project_id, source.content_sha256, suffix)
        if not path.exists():
            raise DataXValidationError("Source snapshot content is missing.")
        connection = duckdb.connect(str(self.store.project_db_path(source.project_id)))
        temp_table = f"tmp_{source.table_name}"
        try:
            connection.execute("BEGIN TRANSACTION")
            connection.execute(f"DROP TABLE IF EXISTS {_quote_ident(temp_table)}")
            if source.file_type == "csv":
                connection.execute(
                    f"CREATE TABLE {_quote_ident(temp_table)} AS SELECT * FROM read_csv_auto(?, header=true, sample_size=-1)",
                    [str(path)],
                )
            elif source.file_type == "parquet":
                connection.execute(
                    f"CREATE TABLE {_quote_ident(temp_table)} AS SELECT * FROM read_parquet(?)",
                    [str(path)],
                )
            else:
                self._load_xlsx(connection, temp_table, path)
            row_count = int(
                connection.execute(f"SELECT COUNT(*) FROM {_quote_ident(temp_table)}").fetchone()[0]
            )
            if row_count > MAX_SOURCE_ROWS:
                raise DataXValidationError("Data source exceeds the 1,000,000 row limit.")
            columns = connection.execute(f"DESCRIBE {_quote_ident(temp_table)}").fetchall()
            if not columns:
                raise DataXValidationError("Data source has no columns.")
            profile = self._profile_table(connection, temp_table, columns, row_count)
            connection.execute(f"DROP TABLE IF EXISTS {_quote_ident(source.table_name)}")
            connection.execute(
                f"ALTER TABLE {_quote_ident(temp_table)} RENAME TO {_quote_ident(source.table_name)}"
            )
            connection.execute("COMMIT")
        except Exception:
            connection.execute("ROLLBACK")
            raise
        finally:
            connection.close()
        return source.model_copy(
            update={
                "status": "ready",
                "row_count": row_count,
                "column_count": len(columns),
                "profile": profile,
                "error": "",
                "updated_at": self.store.now(),
            }
        )

    def _load_xlsx(self, connection: Any, table_name: str, path: Path) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, None)
            if not first:
                raise DataXValidationError("XLSX source is empty.")
            headers = _unique_headers(first)
            sample_rows: list[tuple[Any, ...]] = []
            for _ in range(2000):
                row = next(iterator, None)
                if row is None:
                    break
                sample_rows.append(tuple(row))
            column_types = [
                _infer_xlsx_type(
                    row[index] if index < len(row) else None for row in sample_rows
                )
                for index in range(len(headers))
            ]
            connection.execute(
                f"CREATE TABLE {_quote_ident(table_name)} ({', '.join(f'{_quote_ident(name)} {column_type}' for name, column_type in zip(headers, column_types))})"
            )
            placeholders = ",".join("?" for _ in headers)
            batch: list[list[Any]] = []
            count = 0
            for raw in chain(sample_rows, iterator):
                count += 1
                if count > MAX_SOURCE_ROWS:
                    raise DataXValidationError("Data source exceeds the 1,000,000 row limit.")
                row = [
                    _xlsx_value(
                        raw[index] if index < len(raw) else None,
                        column_types[index],
                    )
                    for index in range(len(headers))
                ]
                batch.append(row)
                if len(batch) >= 1000:
                    connection.executemany(
                        f"INSERT INTO {_quote_ident(table_name)} VALUES ({placeholders})", batch
                    )
                    batch.clear()
            if batch:
                connection.executemany(
                    f"INSERT INTO {_quote_ident(table_name)} VALUES ({placeholders})", batch
                )
        finally:
            workbook.close()

    def _profile_table(
        self, connection: Any, table_name: str, columns: list[Any], row_count: int
    ) -> dict[str, Any]:
        result: list[dict[str, Any]] = []
        for column in columns:
            name = str(column[0])
            data_type = str(column[1])
            quoted = _quote_ident(name)
            nulls, unique = connection.execute(
                f"SELECT COUNT(*) FILTER (WHERE {quoted} IS NULL), COUNT(DISTINCT {quoted}) FROM {_quote_ident(table_name)}"
            ).fetchone()
            item: dict[str, Any] = {
                "name": name,
                "data_type": data_type,
                "null_count": int(nulls),
                "null_rate": round(int(nulls) / row_count, 6) if row_count else 0.0,
                "unique_count": int(unique),
            }
            if any(token in data_type.upper() for token in ("INT", "DOUBLE", "DECIMAL", "FLOAT", "DATE", "TIME")):
                minimum, maximum = connection.execute(
                    f"SELECT MIN({quoted}), MAX({quoted}) FROM {_quote_ident(table_name)}"
                ).fetchone()
                item["min"] = _json_value(minimum)
                item["max"] = _json_value(maximum)
            result.append(item)
        return {"row_count": row_count, "columns": result}

    # Semantic models
    def list_models(self, project_id: str) -> list[SemanticModel]:
        self.get_project(project_id)
        return sorted(
            [item for item in self.store.list("models", SemanticModel) if item.project_id == project_id],
            key=lambda item: (item.created_at, item.model_id),
        )

    def get_model(self, model_id: str) -> SemanticModel:
        return self.store.require("models", model_id, SemanticModel)

    def create_model(
        self,
        project_id: str,
        *,
        name: str,
        description: str = "",
        entities: list[dict[str, Any]],
        joins: list[dict[str, Any]] | None = None,
        fields: list[dict[str, Any]] | None = None,
    ) -> SemanticModel:
        self.get_project(project_id)
        now = self.store.now()
        model = SemanticModel(
            model_id=self.store.new_id("datax_model"),
            project_id=project_id,
            name=name.strip()[:160],
            description=description.strip()[:1000],
            entities=[SemanticEntity.model_validate(item) for item in entities],
            joins=[SemanticJoin.model_validate(item) for item in (joins or [])],
            fields=[ModelField.model_validate(item) for item in (fields or [])],
            created_at=now,
            updated_at=now,
        )
        model = self._validate_model(model)
        return self.store.insert("models", model, "model_id")  # type: ignore[return-value]

    def update_model(self, model_id: str, *, revision: int, patch: dict[str, Any]) -> SemanticModel:
        current = self.get_model(model_id)
        _check_revision(current.revision, revision)
        allowed = {"name", "description", "entities", "joins", "fields"}
        values = current.model_dump()
        for key, value in patch.items():
            if key in allowed and value is not None:
                values[key] = value
        values["revision"] += 1
        values["updated_at"] = self.store.now()
        model = self._validate_model(SemanticModel.model_validate(values))
        return self.store.replace("models", model, "model_id")  # type: ignore[return-value]

    def _validate_model(self, model: SemanticModel) -> SemanticModel:
        if not model.name.strip():
            raise DataXValidationError("Semantic model name is required.")
        entity_ids = [item.entity_id for item in model.entities]
        if len(entity_ids) != len(set(entity_ids)):
            raise DataXValidationError("Semantic entity IDs must be unique.")
        aliases = [item.alias for item in model.entities]
        if len(aliases) != len(set(aliases)) or any(not _is_identifier(item) for item in aliases):
            raise DataXValidationError("Semantic entity aliases must be unique identifiers.")
        sources = {item.source_id: item for item in self.list_sources(model.project_id)}
        for entity in model.entities:
            source = sources.get(entity.source_id)
            if source is None or source.status != "ready":
                raise DataXValidationError("Every semantic entity must reference a ready source.")
        field_names: set[str] = set()
        source_columns = {
            item.source_id: {str(column.get("name")) for column in item.profile.get("columns", [])}
            for item in sources.values()
        }
        for field in model.fields:
            if field.entity_id not in entity_ids:
                raise DataXValidationError("Semantic field references an unknown entity.")
            if not _is_identifier(field.name) or field.name in field_names:
                raise DataXValidationError("Semantic field names must be unique identifiers.")
            entity = next(item for item in model.entities if item.entity_id == field.entity_id)
            if field.source_field not in source_columns.get(entity.source_id, set()):
                raise DataXValidationError(f"Unknown source field: {field.source_field}")
            field_names.add(field.name)
        for join in model.joins:
            if join.left_entity_id not in entity_ids or join.right_entity_id not in entity_ids:
                raise DataXValidationError("Join references an unknown entity.")
            left = next(item for item in model.entities if item.entity_id == join.left_entity_id)
            right = next(item for item in model.entities if item.entity_id == join.right_entity_id)
            if join.left_field not in source_columns.get(left.source_id, set()):
                raise DataXValidationError("Join references an unknown left field.")
            if join.right_field not in source_columns.get(right.source_id, set()):
                raise DataXValidationError("Join references an unknown right field.")
        if len(model.entities) > 1 and len(model.joins) < len(model.entities) - 1:
            raise DataXValidationError("Multi-entity models require explicit joins.")
        return model

    def preview_model(self, model_id: str, *, limit: int = 20) -> dict[str, Any]:
        import duckdb

        model = self.get_model(model_id)
        field_map = self._field_map(model)
        selected = [field for field in model.fields if field.role != "hidden"][:20]
        if not selected:
            raise DataXValidationError("Semantic model has no visible fields.")
        sql = "SELECT " + ", ".join(
            f"{field_map[field.name]} AS {_quote_ident(field.name)}" for field in selected
        )
        sql += " " + self._from_sql(model) + " LIMIT ?"
        connection = duckdb.connect(str(self.store.project_db_path(model.project_id)), read_only=True)
        try:
            cursor = connection.execute(sql, [max(1, min(limit, 100))])
            names = [item[0] for item in cursor.description]
            rows = [_row(names, item) for item in cursor.fetchall()]
        finally:
            connection.close()
        return {"model_id": model_id, "columns": names, "rows": rows, "row_count": len(rows)}

    # Indicators and immutable versions
    def list_indicators(self, project_id: str, *, status: str | None = None) -> list[IndicatorDefinition]:
        self.get_project(project_id)
        items = [
            item
            for item in self.store.list("indicators", IndicatorDefinition)
            if item.project_id == project_id and (status is None or item.status == status)
        ]
        return sorted(items, key=lambda item: (item.name.casefold(), item.indicator_id))

    def get_indicator(self, indicator_id: str) -> IndicatorDefinition:
        return self.store.require("indicators", indicator_id, IndicatorDefinition)

    def get_published_indicator(self, indicator_id: str) -> IndicatorDefinition:
        current = self.get_indicator(indicator_id)
        if current.status == "archived" or current.current_version is None:
            raise DataXNotFoundError("Published Data X indicator was not found.")
        version = self.store.require(
            "indicator_versions",
            f"{indicator_id}:v{current.current_version}",
            IndicatorVersion,
        )
        return IndicatorDefinition.model_validate(version.snapshot)

    def list_published_indicators(self, project_id: str) -> list[IndicatorDefinition]:
        self.get_project(project_id)
        result: list[IndicatorDefinition] = []
        for current in self.list_indicators(project_id):
            if current.status == "archived" or current.current_version is None:
                continue
            result.append(self.get_published_indicator(current.indicator_id))
        return sorted(result, key=lambda item: (item.name.casefold(), item.indicator_id))

    def create_indicator(self, project_id: str, payload: dict[str, Any]) -> IndicatorDefinition:
        self.get_project(project_id)
        now = self.store.now()
        values = dict(payload)
        values.update(
            {
                "indicator_id": self.store.new_id("datax_indicator"),
                "project_id": project_id,
                "status": "draft",
                "revision": 1,
                "current_version": None,
                "created_at": now,
                "updated_at": now,
            }
        )
        item = self.validate_indicator(IndicatorDefinition.model_validate(values))
        self._assert_unique_indicator_code(item)
        return self.store.insert("indicators", item, "indicator_id")  # type: ignore[return-value]

    def update_indicator(
        self, indicator_id: str, *, revision: int, patch: dict[str, Any]
    ) -> IndicatorDefinition:
        current = self.get_indicator(indicator_id)
        _check_revision(current.revision, revision)
        allowed = {
            "model_id", "code", "name", "description", "indicator_type", "aggregation",
            "measure_field", "formula", "default_dimensions", "time_field", "filters", "tags",
        }
        values = current.model_dump()
        for key, value in patch.items():
            if key in allowed and value is not None:
                values[key] = value
        values["revision"] += 1
        values["status"] = "draft"
        values["updated_at"] = self.store.now()
        item = self.validate_indicator(IndicatorDefinition.model_validate(values))
        self._assert_unique_indicator_code(item)
        return self.store.replace("indicators", item, "indicator_id")  # type: ignore[return-value]

    def validate_indicator(self, item: IndicatorDefinition) -> IndicatorDefinition:
        model = self.get_model(item.model_id)
        if model.project_id != item.project_id:
            raise DataXValidationError("Indicator model belongs to another project.")
        if not _is_identifier(item.code):
            raise DataXValidationError("Indicator code must be a safe identifier.")
        fields = {field.name: field for field in model.fields}
        for name in item.default_dimensions:
            if name not in fields or fields[name].role not in {"dimension", "time"}:
                raise DataXValidationError(f"Invalid indicator dimension: {name}")
        if item.time_field and (item.time_field not in fields or fields[item.time_field].role != "time"):
            raise DataXValidationError("Invalid indicator time field.")
        for condition in item.filters:
            if condition.field not in fields:
                raise DataXValidationError(f"Unknown indicator filter field: {condition.field}")
        if item.indicator_type == "basic":
            if item.aggregation is None:
                raise DataXValidationError("Basic indicators require an aggregation.")
            if item.aggregation != "count" and not item.measure_field:
                raise DataXValidationError("This aggregation requires a measure field.")
            if item.measure_field and item.measure_field not in fields:
                raise DataXValidationError("Unknown indicator measure field.")
            item.formula = None
        else:
            if not item.formula:
                raise DataXValidationError("Derived indicators require a formula.")
            dependencies = _formula_dependencies(item.formula)
            published_codes = {
                metric.code
                for metric in self.list_published_indicators(item.project_id)
                if metric.model_id == item.model_id and metric.indicator_id != item.indicator_id
            }
            if not dependencies or not dependencies.issubset(published_codes):
                raise DataXValidationError("Derived formulas may reference only published indicator codes.")
            self._assert_no_formula_cycle(item, dependencies)
            item.aggregation = None
            item.measure_field = None
        return item

    def publish_indicator(self, indicator_id: str, *, revision: int) -> IndicatorDefinition:
        current = self.get_indicator(indicator_id)
        _check_revision(current.revision, revision)
        current = self.validate_indicator(current)
        versions = [
            item
            for item in self.store.list("indicator_versions", IndicatorVersion)
            if item.indicator_id == indicator_id
        ]
        version_number = max((item.version for item in versions), default=0) + 1
        snapshot = current.model_dump(mode="json")
        snapshot["status"] = "published"
        snapshot["current_version"] = version_number
        now = self.store.now()
        version = IndicatorVersion(
            version_id=f"{indicator_id}:v{version_number}",
            indicator_id=indicator_id,
            version=version_number,
            snapshot=snapshot,
            content_sha256=self.store.hash_json(snapshot),
            published_at=now,
        )
        self.store.insert("indicator_versions", version, "version_id")
        updated = current.model_copy(
            update={
                "status": "published",
                "current_version": version_number,
                "revision": current.revision + 1,
                "updated_at": now,
            }
        )
        return self.store.replace("indicators", updated, "indicator_id")  # type: ignore[return-value]

    def archive_indicator(self, indicator_id: str, *, revision: int) -> IndicatorDefinition:
        current = self.get_indicator(indicator_id)
        _check_revision(current.revision, revision)
        updated = current.model_copy(
            update={"status": "archived", "revision": current.revision + 1, "updated_at": self.store.now()}
        )
        return self.store.replace("indicators", updated, "indicator_id")  # type: ignore[return-value]

    def search_indicators(
        self, query: str, *, project_ids: list[str] | None = None, limit: int = 20
    ) -> dict[str, Any]:
        allowed = set(project_ids or [])
        candidates = [
            item
            for project in self.list_projects()
            for item in self.list_published_indicators(project.project_id)
            if not allowed or item.project_id in allowed
        ]
        query_vector = _hash_embedding(query)
        query_tokens = _tokens(query)
        scored: list[tuple[float, IndicatorDefinition]] = []
        for item in candidates:
            text = " ".join([item.code, item.name, item.description, *item.tags])
            lexical = _lexical_score(query_tokens, _tokens(text))
            vector = _cosine(query_vector, _hash_embedding(text))
            scored.append((0.65 * lexical + 0.35 * vector, item))
        scored.sort(key=lambda pair: (-pair[0], pair[1].code))
        selected = [
            {
                "indicator_id": item.indicator_id,
                "project_id": item.project_id,
                "model_id": item.model_id,
                "code": item.code,
                "name": item.name,
                "description": item.description,
                "version": item.current_version,
                "score": round(score, 6),
            }
            for score, item in scored[: max(1, min(limit, 100))]
        ]
        return {
            "items": selected,
            "total": len(selected),
            "retrieval_mode": "local_lexical_hash_vector",
            "warnings": ["External embedding unavailable; using deterministic local hash vectors."],
        }

    def _assert_unique_indicator_code(self, item: IndicatorDefinition) -> None:
        for other in self.store.list("indicators", IndicatorDefinition):
            if (
                other.project_id != item.project_id
                or other.indicator_id == item.indicator_id
                or other.status == "archived"
            ):
                continue
            reserved_codes = {other.code}
            if other.current_version is not None:
                reserved_codes.add(self.get_published_indicator(other.indicator_id).code)
            if item.code in reserved_codes:
                raise DataXConflictError(f"Indicator code already exists: {item.code}")

    def _assert_no_formula_cycle(self, item: IndicatorDefinition, dependencies: set[str]) -> None:
        by_code = {
            metric.code: metric
            for metric in self.list_indicators(item.project_id)
            if metric.model_id == item.model_id
        }
        by_code[item.code] = item

        def visit(code: str, path: set[str]) -> None:
            if code in path:
                raise DataXValidationError("Derived indicator formula contains a cycle.")
            metric = by_code.get(code)
            if metric is None or metric.indicator_type != "derived" or not metric.formula:
                return
            next_path = {*path, code}
            for dependency in _formula_dependencies(metric.formula):
                visit(dependency, next_path)

        for dependency in dependencies:
            visit(dependency, {item.code})

    # Restricted semantic query DSL
    def query(
        self,
        *,
        project_id: str,
        model_id: str,
        indicator_codes: list[str],
        dimensions: list[str] | None = None,
        filters: list[dict[str, Any]] | None = None,
        time_range: dict[str, Any] | None = None,
        order_by: list[dict[str, str]] | None = None,
        limit: int = 100,
        view: str = "table",
        published_only: bool = True,
    ) -> DataXResultArtifact:
        model = self.get_model(model_id)
        if model.project_id != project_id:
            raise DataXValidationError("Semantic model is outside the requested project.")
        if view not in {"kpi", "table", "line", "bar"}:
            raise DataXValidationError("Unsupported Data X result view.")
        row_limit = max(1, min(int(limit), MAX_QUERY_ROWS))
        requested = list(dict.fromkeys(code.strip() for code in indicator_codes if code.strip()))
        if not requested:
            raise DataXValidationError("At least one indicator is required.")
        source_indicators = (
            self.list_published_indicators(project_id)
            if published_only
            else self.list_indicators(project_id)
        )
        available = {
            item.code: item for item in source_indicators if item.model_id == model_id
        }
        missing = [code for code in requested if code not in available]
        if missing:
            raise DataXValidationError("Unknown or unpublished indicators: " + ", ".join(missing))
        dimensions = list(dict.fromkeys(dimensions or []))
        fields = {field.name: field for field in model.fields}
        for name in dimensions:
            if name not in fields or fields[name].role not in {"dimension", "time"}:
                raise DataXValidationError(f"Invalid query dimension: {name}")
        query_filters = [IndicatorFilter.model_validate(item) for item in (filters or [])]
        for condition in query_filters:
            if condition.field not in fields:
                raise DataXValidationError(f"Unknown query filter field: {condition.field}")

        needed_basic: set[str] = set()

        def collect(code: str) -> None:
            metric = available[code]
            if metric.indicator_type == "basic":
                needed_basic.add(code)
                return
            assert metric.formula
            for dependency in _formula_dependencies(metric.formula):
                if dependency not in available:
                    raise DataXValidationError(f"Derived dependency is unavailable: {dependency}")
                collect(dependency)

        for code in requested:
            collect(code)

        combined: dict[tuple[Any, ...], dict[str, Any]] = {}
        started = time.monotonic()
        for code in sorted(needed_basic):
            metric = available[code]
            for row in self._query_basic(
                model,
                metric,
                dimensions,
                query_filters,
                time_range or {},
                row_limit,
            ):
                key = tuple(row.get(name) for name in dimensions)
                current = combined.setdefault(key, {name: row.get(name) for name in dimensions})
                current[code] = row.get(code)
            if time.monotonic() - started > 10:
                raise DataXValidationError("Data X query exceeded the 10 second timeout.")
        if not combined and not dimensions:
            combined[()] = {}
        for row in combined.values():
            for code in requested:
                metric = available[code]
                if metric.indicator_type == "derived":
                    assert metric.formula
                    row[code] = _evaluate_formula(metric.formula, row)
        rows = [{**{name: row.get(name) for name in dimensions}, **{code: row.get(code) for code in requested}} for row in combined.values()]
        for ordering in reversed(order_by or []):
            field = str(ordering.get("field") or "")
            if field not in {*dimensions, *requested}:
                raise DataXValidationError(f"Invalid order field: {field}")
            reverse = str(ordering.get("direction") or "asc").lower() == "desc"
            rows.sort(key=lambda item: (item.get(field) is None, item.get(field)), reverse=reverse)
        truncated = len(rows) > row_limit
        rows = rows[:row_limit]
        artifact = DataXResultArtifact(
            artifact_id=self.store.new_id("datax_result"),
            project_id=project_id,
            model_id=model_id,
            view=view,  # type: ignore[arg-type]
            columns=[*dimensions, *requested],
            rows=[{key: _json_value(value) for key, value in row.items()} for row in rows],
            row_count=len(rows),
            truncated=truncated,
            created_at=self.store.now(),
        )
        self.store.insert("artifacts", artifact, "artifact_id")
        return artifact

    def dimension_members(
        self, model_id: str, field_name: str, *, search: str = "", limit: int = 50
    ) -> list[Any]:
        import duckdb

        model = self.get_model(model_id)
        fields = self._field_map(model)
        expression = fields.get(field_name)
        if expression is None:
            raise DataXValidationError("Unknown semantic field.")
        sql = f"SELECT DISTINCT {expression} AS value {self._from_sql(model)}"
        params: list[Any] = []
        if search:
            sql += f" WHERE CAST({expression} AS VARCHAR) ILIKE ?"
            params.append(f"%{search[:200]}%")
        sql += " ORDER BY value LIMIT ?"
        params.append(max(1, min(limit, 200)))
        connection = duckdb.connect(str(self.store.project_db_path(model.project_id)), read_only=True)
        try:
            return [_json_value(row[0]) for row in connection.execute(sql, params).fetchall()]
        finally:
            connection.close()

    def _query_basic(
        self,
        model: SemanticModel,
        metric: IndicatorDefinition,
        dimensions: list[str],
        request_filters: list[IndicatorFilter],
        time_range: dict[str, Any],
        limit: int,
    ) -> list[dict[str, Any]]:
        import duckdb

        fields = self._field_map(model)
        dimension_sql = [fields[name] for name in dimensions]
        if metric.aggregation == "count" and not metric.measure_field:
            aggregate = "COUNT(*)"
        else:
            if not metric.measure_field:
                raise DataXValidationError("Indicator measure field is required.")
            measure = fields[metric.measure_field]
            function = {
                "sum": "SUM",
                "count": "COUNT",
                "count_distinct": "COUNT(DISTINCT",
                "avg": "AVG",
                "min": "MIN",
                "max": "MAX",
            }[str(metric.aggregation)]
            aggregate = f"{function} {measure})" if function == "COUNT(DISTINCT" else f"{function}({measure})"
        selects = [f"{expression} AS {_quote_ident(name)}" for expression, name in zip(dimension_sql, dimensions)]
        selects.append(f"{aggregate} AS {_quote_ident(metric.code)}")
        conditions: list[str] = []
        params: list[Any] = []
        for item in [*metric.filters, *request_filters]:
            expression = fields[item.field]
            clause, values = _compile_filter(expression, item)
            conditions.append(clause)
            params.extend(values)
        if time_range:
            field_name = str(time_range.get("field") or metric.time_field or "")
            if not field_name or field_name not in fields:
                raise DataXValidationError("A valid time field is required for time_range.")
            if time_range.get("start") not in {None, ""}:
                conditions.append(f"{fields[field_name]} >= ?")
                params.append(time_range["start"])
            if time_range.get("end") not in {None, ""}:
                conditions.append(f"{fields[field_name]} <= ?")
                params.append(time_range["end"])
        sql = f"SELECT {', '.join(selects)} {self._from_sql(model)}"
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        if dimension_sql:
            sql += " GROUP BY " + ", ".join(dimension_sql)
        sql += " LIMIT ?"
        params.append(limit + 1)
        connection = duckdb.connect(str(self.store.project_db_path(model.project_id)), read_only=True)
        try:
            cursor = connection.execute(sql, params)
            names = [item[0] for item in cursor.description]
            return [_row(names, item) for item in cursor.fetchall()]
        finally:
            connection.close()

    def _field_map(self, model: SemanticModel) -> dict[str, str]:
        aliases = {entity.entity_id: entity.alias for entity in model.entities}
        return {
            field.name: f"{_quote_ident(aliases[field.entity_id])}.{_quote_ident(field.source_field)}"
            for field in model.fields
        }

    def _from_sql(self, model: SemanticModel) -> str:
        sources = {source.source_id: source for source in self.list_sources(model.project_id)}
        entities = {entity.entity_id: entity for entity in model.entities}
        first = model.entities[0]
        sql = f"FROM {_quote_ident(sources[first.source_id].table_name)} AS {_quote_ident(first.alias)}"
        joined = {first.entity_id}
        pending = list(model.joins)
        while pending:
            progressed = False
            for join in list(pending):
                if join.left_entity_id in joined and join.right_entity_id not in joined:
                    existing_id, new_id = join.left_entity_id, join.right_entity_id
                    existing_field, new_field = join.left_field, join.right_field
                elif join.right_entity_id in joined and join.left_entity_id not in joined:
                    existing_id, new_id = join.right_entity_id, join.left_entity_id
                    existing_field, new_field = join.right_field, join.left_field
                else:
                    continue
                existing = entities[existing_id]
                new = entities[new_id]
                join_keyword = "INNER JOIN" if join.join_type == "inner" else "LEFT JOIN"
                sql += (
                    f" {join_keyword} {_quote_ident(sources[new.source_id].table_name)} AS {_quote_ident(new.alias)}"
                    f" ON {_quote_ident(existing.alias)}.{_quote_ident(existing_field)}"
                    f" = {_quote_ident(new.alias)}.{_quote_ident(new_field)}"
                )
                joined.add(new_id)
                pending.remove(join)
                progressed = True
            if not progressed:
                raise DataXValidationError("Semantic join graph is disconnected or cyclically ambiguous.")
        if len(joined) != len(model.entities):
            raise DataXValidationError("Semantic join graph does not connect every entity.")
        return sql

    # Approval-gated indicator proposals
    def create_proposal(
        self,
        *,
        project_id: str,
        model_id: str,
        title: str,
        payload: dict[str, Any],
        indicator_id: str | None = None,
        source_xpert_id: str | None = None,
        source_run_id: str | None = None,
        source_goal_id: str | None = None,
        source_handoff_id: str | None = None,
    ) -> IndicatorProposal:
        model = self.get_model(model_id)
        if model.project_id != project_id:
            raise DataXValidationError("Proposal model belongs to another project.")
        proposal_type = "update" if indicator_id else "create"
        if indicator_id:
            current = self.get_indicator(indicator_id)
            if current.project_id != project_id or current.model_id != model_id:
                raise DataXValidationError("Proposal target is outside its configured scope.")
        payload_hash = self.store.hash_json(payload)
        for item in self.store.list("proposals", IndicatorProposal):
            if (
                item.status == "pending"
                and item.project_id == project_id
                and item.model_id == model_id
                and item.indicator_id == indicator_id
                and item.source_run_id == source_run_id
                and self.store.hash_json(item.payload) == payload_hash
            ):
                return item
        now = self.store.now()
        item = IndicatorProposal(
            proposal_id=self.store.new_id("datax_proposal"),
            project_id=project_id,
            model_id=model_id,
            indicator_id=indicator_id,
            proposal_type=proposal_type,
            title=title.strip()[:200] or "Indicator proposal",
            payload=payload,
            source_xpert_id=source_xpert_id,
            source_run_id=source_run_id,
            source_goal_id=source_goal_id,
            source_handoff_id=source_handoff_id,
            created_at=now,
            updated_at=now,
        )
        return self.store.insert("proposals", item, "proposal_id")  # type: ignore[return-value]

    def list_proposals(
        self,
        *,
        project_id: str | None = None,
        status: str | None = None,
        source_xpert_id: str | None = None,
        limit: int = 100,
    ) -> list[IndicatorProposal]:
        items = [
            item
            for item in self.store.list("proposals", IndicatorProposal)
            if (project_id is None or item.project_id == project_id)
            and (status is None or item.status == status)
            and (source_xpert_id is None or item.source_xpert_id == source_xpert_id)
        ]
        return sorted(items, key=lambda item: (-item.updated_at, item.proposal_id))[:limit]

    def get_proposal(self, proposal_id: str) -> IndicatorProposal:
        return self.store.require("proposals", proposal_id, IndicatorProposal)

    def update_proposal(
        self, proposal_id: str, *, revision: int, title: str | None, payload: dict[str, Any] | None
    ) -> IndicatorProposal:
        item = self.get_proposal(proposal_id)
        _check_revision(item.revision, revision)
        if item.status != "pending":
            raise DataXConflictError("Only pending indicator proposals can be edited.")
        values = item.model_dump()
        if title is not None:
            values["title"] = title.strip()[:200]
        if payload is not None:
            values["payload"] = payload
        values["revision"] += 1
        values["updated_at"] = self.store.now()
        updated = IndicatorProposal.model_validate(values)
        return self.store.replace("proposals", updated, "proposal_id")  # type: ignore[return-value]

    def resolve_proposal(
        self,
        proposal_id: str,
        *,
        revision: int,
        action: str,
        operator: str,
        reason: str = "",
    ) -> IndicatorProposal:
        item = self.get_proposal(proposal_id)
        _check_revision(item.revision, revision)
        if item.status != "pending":
            raise DataXConflictError("Indicator proposal has already been resolved.")
        if action == "approve":
            if item.proposal_type == "create":
                indicator = self.create_indicator(item.project_id, item.payload)
            else:
                if not item.indicator_id:
                    raise DataXValidationError("Update proposal has no target indicator.")
                current = self.get_indicator(item.indicator_id)
                indicator = self.update_indicator(
                    current.indicator_id, revision=current.revision, patch=item.payload
                )
            status = "approved"
            reason = f"draft:{indicator.indicator_id}"
        elif action in {"reject", "cancel"}:
            status = "rejected" if action == "reject" else "cancelled"
        else:
            raise DataXValidationError("Unknown proposal action.")
        now = self.store.now()
        updated = item.model_copy(
            update={
                "status": status,
                "revision": item.revision + 1,
                "operator": operator[:120],
                "reason": reason[:1000],
                "updated_at": now,
                "resolved_at": now,
            }
        )
        return self.store.replace("proposals", updated, "proposal_id")  # type: ignore[return-value]

    def capabilities(self) -> dict[str, Any]:
        try:
            import duckdb

            duckdb_version = getattr(duckdb, "__version__", "available")
            duckdb_ready = True
        except Exception:
            duckdb_version = "unavailable"
            duckdb_ready = False
        try:
            import openpyxl

            xlsx_ready = True
            openpyxl_version = getattr(openpyxl, "__version__", "available")
        except Exception:
            xlsx_ready = False
            openpyxl_version = "unavailable"
        return {
            "version": "modelmirror-datax-v1",
            "duckdb_ready": duckdb_ready,
            "duckdb_version": duckdb_version,
            "xlsx_ready": xlsx_ready,
            "openpyxl_version": openpyxl_version,
            "formats": ["csv", "xlsx", "parquet"],
            "aggregations": ["sum", "count", "count_distinct", "avg", "min", "max"],
            "views": ["kpi", "table", "line", "bar"],
            "limits": {
                "source_bytes": MAX_SOURCE_BYTES,
                "source_rows": MAX_SOURCE_ROWS,
                "query_rows": MAX_QUERY_ROWS,
                "query_timeout_seconds": 10,
            },
            "arbitrary_sql": False,
            "embedding_mode": "local_hash_fallback",
        }


def _check_revision(current: int, requested: int) -> None:
    if current != requested:
        raise DataXConflictError(f"Revision conflict: current revision is {current}.")


def _quote_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _is_identifier(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]{0,127}", value or ""))


def _unique_headers(values: Iterable[Any]) -> list[str]:
    result: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(values, start=1):
        base = str(value or f"column_{index}").strip()[:120] or f"column_{index}"
        candidate = base
        suffix = 2
        while candidate in used:
            candidate = f"{base}_{suffix}"
            suffix += 1
        used.add(candidate)
        result.append(candidate)
    return result


def _infer_xlsx_type(values: Iterable[Any]) -> str:
    kinds: set[str] = set()
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool):
            kinds.add("boolean")
        elif isinstance(value, int):
            kinds.add("integer")
        elif isinstance(value, float):
            kinds.add("number")
        elif isinstance(value, (datetime, date)):
            kinds.add("datetime")
        else:
            kinds.add("text")
    if not kinds:
        return "VARCHAR"
    if kinds <= {"integer"}:
        return "BIGINT"
    if kinds <= {"integer", "number"}:
        return "DOUBLE"
    if kinds <= {"boolean"}:
        return "BOOLEAN"
    if kinds <= {"datetime"}:
        return "TIMESTAMP"
    return "VARCHAR"


def _xlsx_value(value: Any, column_type: str) -> Any:
    if value is None:
        return None
    if column_type == "VARCHAR":
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value)
    if column_type == "DOUBLE":
        return float(value)
    if column_type == "BIGINT":
        return int(value)
    if column_type == "BOOLEAN":
        return bool(value)
    if column_type == "TIMESTAMP":
        if isinstance(value, date) and not isinstance(value, datetime):
            return datetime.combine(value, datetime.min.time())
        return value
    return str(value)


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if hasattr(value, "as_integer_ratio") and not isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            pass
    return value


def _row(names: list[str], values: Iterable[Any]) -> dict[str, Any]:
    return {name: _json_value(value) for name, value in zip(names, values)}


def _compile_filter(expression: str, item: IndicatorFilter) -> tuple[str, list[Any]]:
    if item.operator == "eq":
        return f"{expression} = ?", [item.value]
    if item.operator == "ne":
        return f"{expression} <> ?", [item.value]
    if item.operator == "gt":
        return f"{expression} > ?", [item.value]
    if item.operator == "gte":
        return f"{expression} >= ?", [item.value]
    if item.operator == "lt":
        return f"{expression} < ?", [item.value]
    if item.operator == "lte":
        return f"{expression} <= ?", [item.value]
    if item.operator == "contains":
        return f"CAST({expression} AS VARCHAR) ILIKE ?", [f"%{str(item.value)[:500]}%"]
    if item.operator == "in":
        if not isinstance(item.value, list) or not item.value or len(item.value) > 100:
            raise DataXValidationError("The in operator requires between 1 and 100 values.")
        return f"{expression} IN ({','.join('?' for _ in item.value)})", list(item.value)
    raise DataXValidationError("Unsupported filter operator.")


def _formula_dependencies(formula: str) -> set[str]:
    if len(formula) > 1000:
        raise DataXValidationError("Derived formula is too long.")
    try:
        tree = ast.parse(formula, mode="eval")
    except SyntaxError as exc:
        raise DataXValidationError("Derived formula is invalid.") from exc
    for node in ast.walk(tree):
        if not isinstance(node, ALLOWED_FORMULA_NODES):
            raise DataXValidationError("Derived formula contains an unsupported expression.")
        if isinstance(node, ast.Constant) and not isinstance(node.value, (int, float)):
            raise DataXValidationError("Derived formula constants must be numeric.")
    return {node.id for node in ast.walk(tree) if isinstance(node, ast.Name)}


def _evaluate_formula(formula: str, values: dict[str, Any]) -> float | None:
    tree = ast.parse(formula, mode="eval")

    def evaluate(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return evaluate(node.body)
        if isinstance(node, ast.Constant):
            return float(node.value)
        if isinstance(node, ast.Name):
            value = values.get(node.id)
            return float(value or 0)
        if isinstance(node, ast.UnaryOp):
            value = evaluate(node.operand)
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = evaluate(node.left), evaluate(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                if right == 0:
                    raise DataXValidationError("Derived indicator division by zero.")
                return left / right
        raise DataXValidationError("Unsupported derived formula expression.")

    value = evaluate(tree)
    return None if math.isnan(value) or math.isinf(value) else value


def _tokens(value: str) -> list[str]:
    normalized = value.casefold()
    words = re.findall(r"[a-z0-9_]+|[\u3400-\u9fff]", normalized)
    cjk = [char for char in words if len(char) == 1 and "\u3400" <= char <= "\u9fff"]
    bigrams = ["".join(cjk[index : index + 2]) for index in range(max(0, len(cjk) - 1))]
    return words + bigrams


def _hash_embedding(value: str, dimensions: int = 128) -> list[float]:
    vector = [0.0] * dimensions
    for token in _tokens(value):
        digest = hashlib.sha256(token.encode("utf-8")).digest()
        index = int.from_bytes(digest[:4], "big") % dimensions
        vector[index] += -1.0 if digest[4] & 1 else 1.0
    norm = math.sqrt(sum(item * item for item in vector)) or 1.0
    return [item / norm for item in vector]


def _cosine(left: list[float], right: list[float]) -> float:
    return sum(a * b for a, b in zip(left, right))


def _lexical_score(query: list[str], candidate: list[str]) -> float:
    if not query:
        return 0.0
    candidate_set = set(candidate)
    return sum(1 for token in set(query) if token in candidate_set) / len(set(query))


def _safe_error(exc: Exception) -> str:
    message = " ".join(str(exc).split())[:500]
    message = re.sub(r"(?i)(api[_-]?key|token|secret)\s*[:=]\s*\S+", r"\1=[redacted]", message)
    return message or exc.__class__.__name__
